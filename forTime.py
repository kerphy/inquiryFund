#!/usr/bin/env python
# _*_ coding:utf-8 _*_
from openpyxl import load_workbook
from datetime import datetime
import time,requests
import smtplib
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.header import Header
from email.mime.multipart import MIMEMultipart
import os,platform
from git import Repo

#邮件数据
mail = {'me':'309506489@qq.com','host':'smtp.qq.com','port':'465','pw':'qisgaudsxbunbhef'}
# 多个收件人用list,单个收件人字符串
accpter = ['1045033116@qq.com','18612404428@163.com','likelin_work@163.com','Zhangdan0525@outlook.com','ericforever@dingtalk.com']
# accpter = ['1045033116@qq.com']
#每次循环等待间隔时间,默认60秒程序唤醒一次
waitTime = 60
#需要定时监测的时间点
timeList = ['11:30','14:45']
#基金代码
fundCode =['161725','000311','110022','161616','486001','001629','002611','502048']
#路径
base_path = str(os.getcwd()).replace("\\", "/")
excel_path = base_path+'/myfundStatistics.xlsx'

def runTaskRegularTime():
	while True:
		str_time_now = datetime.now().strftime('%H:%M')
		finish_time = datetime.now().strftime('%H')
		if int(finish_time)<15:#3点后结束
			if str(str_time_now) in timeList:#按照设定好的时间对比
			# if 1>0:#调试用
				sendEmail(doData(),False)
				time.sleep(waitTime)
			else:
				taskWait(str_time_now)
		else:
			sys = platform.system()
			if sys == "Windows":
				waitTime2 = 300
				# waitTime2 = 1
			else :
				waitTime2 = 120
			print('闭市了\n等%s分钟后再获取数据'%(waitTime2/60))
			time.sleep(waitTime2)
			matrix = doData()
			write_excel(matrix)
			sendEmail(matrix,True)
			pushExcel()
			break

#把基金返回的数据存到二维数组里
def doData():
	matrix = []
	for n in range(len(fundCode)):
		matrix.append(inquiryRate(fundCode[n]))
	# print(matrix[0])
	return matrix

def taskWait(currentTime):
	print('current time：' + str(currentTime))
	print('process will wakes up every %s seconds because it does not arrive at the specified time' % (waitTime))
	print ('')
	print('')
	time.sleep(waitTime)

#调接口查询
def inquiryRate(shortNumber):
	try:
		url = (" http://fundmobapi.eastmoney.com/FundMApi/FundVarietieValuationDetail.ashx?version=5.3.0&plat=Android&appType=ttjj&FCODE=%s&deviceid=e19655bcbef4c4b362d908f8bcbdd67f||946596830144568&product=EFund&MobileKey=e19655bcbef4c4b362d908f8bcbdd67f||946596830144568" % shortNumber)
		headers = {'Connection': 'close', }
		r = requests.get(url, params=None,headers = headers)
		# print(r.json()["Expansion"])
		today = r.json()["Expansion"]["GZTIME"][0:10]#只截取到日期，不要时间
		fcode = r.json()["Expansion"]["FCODE"]
		shortname = r.json()["Expansion"]["SHORTNAME"]
		yn = r.json()["Expansion"]["DWJZ"]
		tn = r.json()["Expansion"]["GZ"]
		upordown = r.json()["Expansion"]["GSZZL"]
		# print("_____________________")
		ex_content = [today, fcode, shortname, yn, tn, upordown] 
		return ex_content
	except Exception as e:
		print('接口有问题'+str(e))
	return

def sendEmail(list,attach):
	msg = MIMEMultipart()
	msg['from'] = mail['me']
	msg['to'] = ','.join(accpter)
	msg['subject'] = Header("%s基金数据,白酒%s"%(datetime.now().strftime('%H:%M'),list[0][5]), 'utf-8')#当前时间与第一个基金的涨幅
	list1= []
	for n in range(len(list)):
		list1.append('    '.join(list[n]))
	mailTitle = '日期              基金代码         基金名                开盘净值   目前净值     涨跌幅度%\n'
	mailContent = '\n'.join(list1)
	mailBody = mailTitle+mailContent
	print(mailBody)
	puretext = MIMEText(mailBody, 'plain', 'utf-8')
	msg.attach(puretext)
	if attach:
		excel_msg = MIMEApplication(open(excel_path, 'rb').read())
		excel_msg.add_header('Content-Disposition', 'attachment', filename='statistics.xlsx')
		msg.attach(excel_msg)
	try:
		server = smtplib.SMTP_SSL(mail['host'], mail['port'])
		server.login(mail['me'], mail['pw'])
		server.sendmail(mail['me'], accpter, msg.as_string())
		server.quit()
	except Exception as e:
		print(str(e))
	return

#读取项目的excel，修改添加数据后保存
def write_excel(list):
	wb = load_workbook(excel_path)
	sheet = wb['Sheet1']
	n_coordinate = 0
	for row in sheet.iter_rows():#找到空白行号
		for cell in row:
			# print(cell.coordinate, cell.value)
			if cell.value is not None:
				n_coordinate = cell.row+1
	# print(n_coordinate)
	for n in range(len(list)):#通过遍历给cell赋值
		for m in range(len(list[0])):
			sheet.cell(n_coordinate+n,m+1).value=list[n][m]
	wb.save(excel_path)

#push到github
def	pushExcel():
	repo = Repo(base_path)
	remote = repo.remote()
	repo.git.checkout()
	index = repo.index
	index.add(['myfundStatistics.xlsx'])
	index.commit('pushed by GitPython')
	remote.push()

if __name__ == '__main__':
	runTaskRegularTime()